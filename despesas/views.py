from django.shortcuts import render, redirect
from django.http import HttpResponse
import pandas as pd
import json
from .models import Despesa, ArquivoResultado, Extra, ConfigProduto, ConfigOrigem
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, date



@login_required(login_url='/login/')
def listar_despesas(request):
    despesas = Despesa.objects.all() 
    
    # NOVO: Calcula a soma de todos os valores da coluna 'valor'
    total_geral = despesas.aggregate(total=Sum('valor'))['total']
    if total_geral is None:
        total_geral = 0  # Se não tiver despesas, o total é 0
        
    return render(request, 'lista_despesas.html', {
        'despesas': despesas,
        'total_geral': total_geral # NOVO: Envia o total para o HTML
    })

@login_required(login_url='/login/')
def criar_despesa(request):
    if request.method == 'POST':
        data = request.POST.get('data')
        valor = request.POST.get('valor')
        origem = request.POST.get('origem')
        categoria = request.POST.get('categoria')
        subcategoria = request.POST.get('subcategoria')
        numero = request.POST.get('numero')
        info = request.POST.get('info')
        descricao = request.POST.get('descricao')
        
        Despesa.objects.create(
            data=data, valor=valor, origem=origem, categoria=categoria,
            subcategoria=subcategoria, numero=numero, info=info, descricao=descricao
        )
        # Pega o valor do botão que foi clicado
        acao = request.POST.get('acao')
        
        if acao == 'salvar_novo':
            # Se clicou no azul, recarrega a própria página de novo lançamento
            return redirect('criar_despesa')
        else:
            # Se clicou no verde (ou apertou Enter), volta para a página inicial
            return redirect('home')
    
   
    
    return render(request, 'novo_lancamento.html')

@login_required(login_url='/login/')
def excluir_despesa(request, id):
    despesa = Despesa.objects.get(id=id)
    despesa.delete()
    return redirect('home')

@login_required(login_url='/login/')
def editar_despesa(request, id):
    # 1. Vai no banco de dados e busca a despesa exata com aquele ID
    despesa = Despesa.objects.get(id=id)
    
    # 2. Se o usuário acabou de clicar no botão "Salvar Alterações" (POST)
    if request.method == 'POST':
        despesa.data = request.POST.get('data')
        despesa.valor = request.POST.get('valor')
        despesa.origem = request.POST.get('origem')
        despesa.categoria = request.POST.get('categoria')
        despesa.subcategoria = request.POST.get('subcategoria')
        despesa.numero = request.POST.get('numero')
        despesa.info = request.POST.get('info')
        despesa.descricao = request.POST.get('descricao')
        
        # Salva as alterações no banco de dados (UPDATE)
        despesa.save()
        
        # Volta para a lista
        return redirect('home')
    
    # 3. Se for apenas entrar na página (GET), manda o HTML com os dados dessa despesa
    return render(request, 'editar_lancamento.html', {'despesa': despesa})


def tela_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            # Se errar a senha, recarrega a página com erro
            return render(request, 'login.html', {'erro': 'Usuário ou senha incorretos!'})
    return render(request, 'login.html')

def deslogar(request):
    logout(request)
    return redirect('tela_login') 

@login_required(login_url='/login/')
def tela_exportar(request):
    return render(request, 'exportar_excel.html')

@login_required(login_url='/login/')
def exportar_excel(request):
    # 1. Puxa todos os dados do banco
    dados = Despesa.objects.all().values('data', 'origem', 'categoria', 'subcategoria', 'numero', 'valor', 'info', 'descricao')
    
    # 2. Transforma em uma tabela do Pandas
    df = pd.DataFrame(dados)
    
    # 3. Prepara a resposta do navegador como um arquivo de download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="minhas_despesas.xlsx"'
    
    # 4. Salva o Pandas dentro do arquivo do navegador
    df.to_excel(response, index=False, engine='openpyxl')
    
    # 5. Entrega o arquivo
    return response

@login_required(login_url='/login/')
def resultado_page(request):
    arquivos_salvos = ArquivoResultado.objects.all().order_by('-data_upload')
    tabela_html = ""
    arquivo_selecionado = None
    abas_disponiveis = []
    aba_selecionada = None

    if request.method == 'POST' and 'arquivo_excel' in request.FILES:
        arquivo = request.FILES['arquivo_excel']
        abas_escolhidas = request.POST.get('abas_liberadas', '')
        ArquivoResultado.objects.create(nome=arquivo.name, arquivo=arquivo, abas_liberadas=abas_escolhidas)
        return redirect('resultado_page')

    arquivo_id = request.GET.get('arquivo_id')
    if arquivo_id:
        try:
            arquivo_obj = ArquivoResultado.objects.get(id=arquivo_id)
            arquivo_selecionado = arquivo_obj
            
            # MÁGICA: Usamos o OpenPyXL para ler o arquivo
            wb = openpyxl.load_workbook(arquivo_obj.arquivo.path, data_only=True)
            
            if arquivo_obj.abas_liberadas:
                abas_disponiveis = [aba.strip() for aba in arquivo_obj.abas_liberadas.split(',') if aba.strip()]
            else:
                abas_disponiveis = wb.sheetnames

            aba_requisitada = request.GET.get('aba')
            
            if aba_requisitada in abas_disponiveis:
                aba_selecionada = aba_requisitada
            else:
                aba_selecionada = abas_disponiveis[0]

            # Pega a aba escolhida
            ws = wb[aba_selecionada]
            
            # Função auxiliar para descobrir a cor real (o Excel é confuso com cores)
            def pegar_cor(cor_obj):
                if not cor_obj or cor_obj.type == 'indexed': return None
                if cor_obj.rgb == '00000000': return None
                rgb = str(cor_obj.rgb)
                if len(rgb) == 8 and rgb.startswith('00'): return f"#{rgb[2:]}"
                return f"#{rgb}"

                    # NOVO: A lista NEGRA. O que estiver aqui, NÃO vira dinheiro. O resto, vira!
            # Adicione mais palavras aqui se precisar ignorar mais colunas
            palavras_que_NAOviram_dinheiro = [
                'quantidade', 'qtd', 'mês', 'mes', 'month', 'dia', 'day', 
                'id', 'código', 'codigo', 'code', 'número', 'numero', 'year', 'ano',
                'unidade', 'item', 'status', 'pago', 'categoria', 'subcategoria','Nights'
            ]
            
            # Pega os cabeçalhos da primeira linha para saber o nome das colunas
            cabecalhos = []
            for cell in ws[1]:
                cabecalhos.append(str(cell.value).lower() if cell.value else "")

            # Começa a construir a tabela HTML com estilos
            html_builder = ['<table style="border-collapse: collapse; font-family: Calibri, sans-serif;">']
            
            for linha_idx, row in enumerate(ws.iter_rows()):
                html_builder.append('<tr>')
                
                for coluna_idx, cell in enumerate(row):
                    estilos_css = "border: 1px solid #d4d4d4; padding: 5px 10px;"
                    valor = cell.value if cell.value is not None else ""

                    # NOVO: Se o Excel identificou como Data, formata para Mês/Ano
                    if isinstance(valor, (datetime, date)):
                        valor = valor.strftime('%m/%Y')
                    
                    # Verifica se é exatamente uma string vazia (""). 
                    if valor == "":
                        estilos_css = "border: none; padding: 0;"
                    
                    # 1. Fundo (Background)
                    if cell.fill and cell.fill.fgColor:
                        cor_fundo = pegar_cor(cell.fill.fgColor)
                        if cor_fundo:
                            estilos_css += f"background-color: {cor_fundo};"
                    
                    # 2. Fonte (Cor, Tamanho, Negrito)
                    if cell.font:
                        if cell.font.color:
                            cor_letra = pegar_cor(cell.font.color)
                            if cor_letra:
                                estilos_css += f"color: {cor_letra};"
                        if cell.font.size:
                            try:
                                tamanho = cell.font.size.pt
                            except AttributeError:
                                tamanho = float(cell.font.size)
                            tamanho_px = int(tamanho * 1.33)
                            estilos_css += f"font-size: {tamanho_px}px;"
                        if cell.font.bold:
                            estilos_css += "font-weight: bold;"

                    # 3. A LÓGICA INTELIGENTE (Dinheiro, Porcentagem, Vizinho ou Normal)
                    if isinstance(valor, (int, float)):
                        nome_da_coluna = cabecalhos[coluna_idx] if coluna_idx < len(cabecalhos) else ""
                        nao_e_dinheiro = False
                        
                        # REGRA A: É PORCENTAGEM nativa do Excel?
                        if cell.number_format and '%' in str(cell.number_format):
                            valor = f"{valor * 100:.2f}%"
                            nao_e_dinheiro = True
                        
                        # REGRA B: A célula AO LADO esquerda diz "Nights"?
                        elif coluna_idx > 0:
                            celula_esquerda = row[coluna_idx - 1]
                            if celula_esquerda.value and isinstance(celula_esquerda.value, str):
                                if 'nights' in celula_esquerda.value.lower():
                                    valor = f"{valor:.0f}"
                                    nao_e_dinheiro = True
                                    
                        # REGRA C: Está na lista negra pelo cabeçalho da coluna?
                        elif any(palavra in nome_da_coluna for palavra in palavras_que_NAOviram_dinheiro):
                            valor = f"{valor:.2f}"
                            nao_e_dinheiro = True
                            
                        # REGRA D: Se não caiu em nenhuma regra acima, vira Dinheiro!
                        if not nao_e_dinheiro:
                            valor = f"${valor:,.2f}"        
                    
                    html_builder.append(f'<td style="{estilos_css}">{valor}</td>')
                html_builder.append('</tr>')
            
            html_builder.append('</table>')
            tabela_html = "".join(html_builder)    
            
        except Exception as e:
            tabela_html = f"<p style='color:red;'>Erro ao processar o estilo da planilha: {e}</p>"

    return render(request, 'resultado.html', {
        'arquivos_salvos': arquivos_salvos,
        'tabela_html': tabela_html,
        'arquivo_selecionado': arquivo_selecionado,
        'abas_disponiveis': abas_disponiveis,
        'aba_selecionada': aba_selecionada
    })

@login_required(login_url='/login/')
def listar_extras(request):
    extras = Extra.objects.all().order_by('-data')
    return render(request, 'lista_extras.html', {'extras': extras})

@login_required(login_url='/login/')
def novo_extra(request):
    produtos_json = "[]"
    origens_json = "[]"

    # Tenta ler o último Excel de Produtos enviado
    config_prod = ConfigProduto.objects.all().order_by('-id').first()
    if config_prod:
        try:
            df_prod = pd.read_excel(config_prod.arquivo.path)
            # Transforma em formato de chave-valor: {"Produto A": 10.5, "Produto B": 5.0}
            dict_produtos = dict(zip(df_prod.iloc[:, 0], df_prod.iloc[:, 1]))
            produtos_json = json.dumps(dict_produtos)
        except:
            pass

    # Tenta ler o último Excel de Origens enviado
    config_orig = ConfigOrigem.objects.all().order_by('-id').first()
    if config_orig:
        try:
            df_orig = pd.read_excel(config_orig.arquivo.path)
            origens_json = json.dumps(df_orig.iloc[:, 0].tolist())
        except:
            pass

    if request.method == 'POST':
        # 1. Pega os valores e garante que viram números (float)
        amount = float(request.POST.get('amount') or 0)
        unitary_value = float(request.POST.get('unitary_value') or 0)
        
        # 2. Faz a conta mágica
        total_calculado = amount * unitary_value

        # 3. Salva no banco de dados já com o total calculado
        Extra.objects.create(
            data=request.POST.get('data'),
            product=request.POST.get('product'),
            amount=amount,
            unitary_value=unitary_value,
            total=total_calculado,
            origin=request.POST.get('origin')
        )
        return redirect('listar_extras')

  
    return render(request, 'novo_extra.html', {'produtos_json': produtos_json, 'origens_json': origens_json})

@login_required(login_url='/login/')
def upload_config_extras(request):
    if request.method == 'POST':
        if 'arquivo_produtos' in request.FILES:
            ConfigProduto.objects.create(arquivo=request.FILES['arquivo_produtos'])
        if 'arquivo_origens' in request.FILES:
            ConfigOrigem.objects.create(arquivo=request.FILES['arquivo_origens'])
    return redirect('novo_extra')


def excluir_resultado(request, id):
    # Vai no banco de dados e busca o arquivo
    arquivo = ArquivoResultado.objects.get(id=id)
    
    # O Django deleta o arquivo físico da pasta media/ E o registro do banco ao mesmo tempo!
    arquivo.delete()
    
    # Volta para a página de resultados
    return redirect('resultado_page')